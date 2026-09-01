#!/usr/bin/env python3
"""One-shot upload of the copywriter's filled translation workbook.

Reads the 'Strings' sheet, takes every row whose Arabic column is filled,
groups rows by source document, and writes one `translation` document per
source (_id `ar--<docId>`, plus `ar--ui` for the fixed chrome strings).
Rows left empty keep their English on the site. Re-running replaces the
translation docs wholesale, so a corrected sheet is just another run.

Usage:
    python3 scripts/i18n/import_ar.py path/to/trc-arabic-translation.xlsx
    (then, from studio-trc/):
    npx sanity dataset import ../import/i18n/translations.ndjson production --replace
"""
import hashlib
import json
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook

if len(sys.argv) < 2:
    sys.exit("usage: import_ar.py <filled-workbook.xlsx>")

wb = load_workbook(sys.argv[1], data_only=True)
ws = wb["Strings"]

docs = defaultdict(list)
skipped = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ref, _where, _field, _en, ar, *_ = row + (None,) * (6 - len(row))
    if not ref or not isinstance(ar, str) or not ar.strip():
        skipped += 1
        continue
    doc_id, path = str(ref).split("|", 1)
    key = "s" + hashlib.md5(path.encode()).hexdigest()[:12]
    docs[doc_id].append({"_key": key, "_type": "object", "path": path, "value": ar.strip()})

out_path = "import/i18n/translations.ndjson"
os.makedirs("import/i18n", exist_ok=True)
with open(out_path, "w") as f:
    for doc_id, strings in docs.items():
        f.write(json.dumps({
            "_id": f"ar--{doc_id.replace('.', '-')}",
            "_type": "translation",
            "lang": "ar",
            "source": doc_id,
            "strings": strings,
        }, ensure_ascii=False) + "\n")

total = sum(len(v) for v in docs.values())
print(f"{total} Arabic strings across {len(docs)} documents -> {out_path}")
print(f"({skipped} rows left English)")
print("\nNow run:\n  cd studio-trc && npx sanity dataset import ../import/i18n/translations.ndjson production --replace")
