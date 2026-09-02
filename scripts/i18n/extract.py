#!/usr/bin/env python3
"""Extract every user-facing English string from Sanity into a translation
workbook. Each row carries a stable Ref (docId|fieldPath) so the filled
sheet can be imported back in one shot by import_ar.py — the copywriter
only ever touches the Arabic column."""
import json
import urllib.request
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT = "nvmppjc2"
DATASET = "production"
API = f"https://{PROJECT}.api.sanity.io/v2026-08-01/data/query/{DATASET}"


def q(groq):
    url = f"{API}?query={urllib.parse.quote(groq)}&perspective=published"
    with urllib.request.urlopen(url) as r:
        return json.load(r)["result"]


rows = []  # (docId, path, where, english, note)


def add(doc_id, path, where, text, note=""):
    if isinstance(text, str) and text.strip():
        rows.append((doc_id, path, where, text.strip(), note))


def title_pair(doc_id, base, where, t):
    if isinstance(t, dict):
        add(doc_id, f"{base}.line1", where, t.get("line1"))
        add(doc_id, f"{base}.line2", where, t.get("line2"))
    elif isinstance(t, str):
        add(doc_id, base, where, t)


# ---- home destinations ----
for d in q('*[_type=="destination"]|order(order asc)'):
    i = d["_id"]
    w = f"Home · {d.get('navLabel', i)}"
    add(i, "navLabel", w, d.get("navLabel"), "top navigation word")
    add(i, "eyebrow", w, d.get("eyebrow"))
    title_pair(i, "title", w, d.get("title"))
    add(i, "copy", w, d.get("copy"))
    add(i, "statusLabel", w, d.get("statusLabel"), "'Approaching …' status text")
    add(i, "season", w, d.get("season"))
    add(i, "interest", w, d.get("interest"), "enquiry-form interest option")
    for n, h in enumerate(d.get("highlights") or []):
        add(i, f"highlights[{n}]", w, h, "small tag")

# ---- regions (stops + catalog labels) ----
for r in q('*[_type=="region"]'):
    i = r["_id"]
    rw = f"Region · {r['slug']['current']}"
    title_pair(i, "title", rw, r.get("title"))
    add(i, "intro", rw, r.get("intro"))
    for s in r.get("stops") or []:
        k = s["_key"]
        w = f"{rw} · {s.get('country','')}"
        add(i, f"stops[{k}].country", w, s.get("country"), "country name as shown")
        add(i, f"stops[{k}].eyebrow", w, s.get("eyebrow"))
        title_pair(i, f"stops[{k}].title", w, s.get("title"))
        add(i, f"stops[{k}].copy", w, s.get("copy"))
        add(i, f"stops[{k}].season", w, s.get("season"))
        for n, h in enumerate(s.get("highlights") or []):
            add(i, f"stops[{k}].highlights[{n}]", w, h, "small tag")
    for g in r.get("catalog") or []:
        add(i, f"catalog[{g['_key']}].label", rw, g.get("label"), "country group label")

# ---- stays ----
for s in q('*[_type=="stay"]|order(name asc)'):
    i = s["_id"]
    w = f"Stay · {s.get('name', i)}"
    add(i, "name", w, s.get("name"), "brand name — leave Arabic EMPTY to keep as-is")
    add(i, "location", w, s.get("location"))
    add(i, "description", w, s.get("description"))
    for n, h in enumerate(s.get("highlights") or []):
        add(i, f"highlights[{n}]", w, h, "small tag")
    for f in s.get("facts") or []:
        k = f["_key"]
        add(i, f"facts[{k}].label", w, f.get("label"), "facts-ledger label")
        add(i, f"facts[{k}].value", w, f.get("value"), "facts-ledger value")
    for a in s.get("assets") or []:
        add(i, f"assets[{a['_key']}].title", w, a.get("title"), "brochure download title")

# ---- country pages ----
for p in q('*[_type=="countryPage"]'):
    i = p["_id"]
    w = f"Page · {p.get('country', i)}"
    add(i, "country", w, p.get("country"), "hero cut-out word + headings")
    add(i, "tagline", w, p.get("tagline"))
    add(i, "priceLine", w, p.get("priceLine"))
    add(i, "season", w, p.get("season"))
    quote = p.get("quote") or {}
    add(i, "quote.text", w, quote.get("text"))
    add(i, "quote.attribution", w, quote.get("attribution"))
    for c in p.get("chapters") or []:
        k = c["_key"]
        cw = f"{w} · chapter {c.get('navLabel','')}"
        add(i, f"chapters[{k}].navLabel", cw, c.get("navLabel"), "section chip")
        add(i, f"chapters[{k}].eyebrow", cw, c.get("eyebrow"))
        title_pair(i, f"chapters[{k}].title", cw, c.get("title"))
        for n, para in enumerate(c.get("paragraphs") or []):
            add(i, f"chapters[{k}].paragraphs[{n}]", cw, para)
    for d in p.get("days") or []:
        k = d["_key"]
        dw = f"{w} · signature {d.get('title','')}"
        add(i, f"days[{k}].title", dw, d.get("title"))
        add(i, f"days[{k}].copy", dw, d.get("copy"))
        for n, det in enumerate(d.get("details") or []):
            add(i, f"days[{k}].details[{n}]", dw, det, "small tag")
    for e in p.get("essentials") or []:
        k = e["_key"]
        ew = f"{w} · essentials {e.get('title','')}"
        add(i, f"essentials[{k}].title", ew, e.get("title"))
        add(i, f"essentials[{k}].copy", ew, e.get("copy"))
        for pt in e.get("points") or []:
            add(i, f"essentials[{k}].points[{pt['_key']}].label", ew, pt.get("label"))
            add(i, f"essentials[{k}].points[{pt['_key']}].value", ew, pt.get("value"))

# ---- fixed UI chrome (lives in code, keyed for the frontend) ----
UI = [
    ("nav.about", "Top navigation", "About"),
    ("nav.enquire", "Top navigation", "Enquire"),
    ("hero.brand.the", "Home hero", "The Retreat"),
    ("hero.brand.collection", "Home hero", "Collection"),
    ("hero.kicker", "Home hero", "An itinerary, mapped — scroll to follow the route"),
    ("hero.scroll", "Scroll hints", "Scroll"),
    ("hero.scrollRoute", "Country page hero", "Scroll to fly the route"),
    ("loader.preparing", "Loading screen", "Preparing route"),
    ("status.approaching", "Status bar", "TRC 001 · Approaching"),
    ("stop.bestSeason", "Every stop / page", "Best season"),
    ("stop.enquireRoute", "Stop CTA", "Enquire about this route"),
    ("strip.choose", "Asia selector", "Choose your route"),
    ("strip.heading", "Asia selector", "{n} countries, {n} ways in"),
    ("strip.explore", "Country cards", "Explore {country} →"),
    ("board.choose", "Mountains selector", "Choose your altitude"),
    ("board.heading", "Mountains selector", "{n} countries, from the Alps to the ice"),
    ("fan.kicker", "Coast selector", "Postcards from the coast"),
    ("fan.heading", "Coast selector", "{n} countries, wish you were here"),
    ("fan.hint", "Coast selector", "scroll to deal the next card · click the front card to travel"),
    ("caravan.kicker", "Desert selector", "The caravan route"),
    ("caravan.heading", "Desert selector", "{n} countries, waypoint by waypoint"),
    ("page.home", "Country page breadcrumb", "Home"),
    ("page.overview", "Country page nav", "Overview"),
    ("page.signatures", "Country page nav", "Signatures"),
    ("page.essentials", "Country page nav", "Essentials"),
    ("page.gallery", "Country page nav", "Gallery"),
    ("page.theStays", "Country page nav", "The stays"),
    ("page.framed", "Gallery heading", "{country}, framed"),
    ("page.galleryHint", "Gallery", "frames · hover to play · tap to open"),
    ("page.whereStay", "Stays section", "Where you'll stay"),
    ("page.staysIn", "Stays section", "The stays in {country}"),
    ("page.staysCount", "Stays section", "stays"),
    ("page.otherRoutes", "Stays section", "Other routes in"),
    ("page.speakToUs", "Chapter CTA", "Speak to us about {country}"),
    ("page.theSignatures", "Signatures section", "The signatures"),
    ("page.signatureBySignature", "Signatures heading", "{country}, signature by signature"),
    ("page.notItinerary", "Signatures section", "Not an itinerary — the marks of the Maison, one signature at a time."),
    ("page.beforeYouPack", "Essentials heading", "Before you pack"),
    ("page.practicalSide", "Essentials", "The practical side of {country}, one sheet at a time."),
    ("dossier.title", "Stay dossier", "Stay dossier"),
    ("dossier.open", "Stay cards", "Open the dossier"),
    ("dossier.enquire", "Stay cards", "Enquire"),
    ("dossier.closerLook", "Stay dossier", "A closer look"),
    ("dossier.brochures", "Stay dossier", "Brochures"),
    ("dossier.download", "Stay dossier", "Download"),
    ("dossier.enquireStay", "Stay dossier", "Enquire about this stay"),
    ("dossier.film", "Media badges", "Film"),
    ("dossier.still", "Media badges", "Still"),
    ("gen.addressNav", "Generated pages", "The Address"),
    ("gen.addressTitle1", "Generated pages", "Rooms worth the"),
    ("gen.addressTitle2", "Generated pages", "flight over"),
    ("gen.chapter1Para", "Generated pages", "Every stay in {country} is visited, vetted, and arranged directly — the route, the rooms, and the hours in between are built around the traveller, not a package."),
    ("gen.addressPara2", "Generated pages", "{names} — each visited, vetted, and held to the same standard."),
    ("gen.addressFallback", "Generated pages", "From {location} outward, the collection keeps only the addresses we would send our own families to."),
    ("gen.quote", "Generated pages", "{country} rewards the traveller who arrives with time to spend, not a list to finish."),
    ("gen.attribution", "Generated pages", "Field notes — The Retreat Collection"),
    ("gen.dayFallback", "Generated pages", "{name}, {location} — arranged directly, with the route and the rooms built around the traveller."),
    ("gen.priceLine", "Generated pages", "Private routings · rates on request"),
    ("gen.gettingThere", "Generated pages", "Getting there"),
    ("gen.gettingThereCopy", "Generated pages", "Routed privately from arrival onward — the route into {country} is arranged end to end before departure."),
    ("gen.arrival", "Generated pages", "Arrival"),
    ("gen.metAirside", "Generated pages", "Met airside"),
    ("gen.transfer", "Generated pages", "Transfer"),
    ("gen.private", "Generated pages", "Private"),
    ("gen.checkin", "Generated pages", "Check-in"),
    ("gen.handled", "Generated pages", "Handled"),
    ("gen.whenToGo", "Generated pages", "When to go"),
    ("gen.whenCopy", "Generated pages", "{season} is the season we plan around; the exact week is chosen with you."),
    ("gen.bookedAhead", "Generated pages", "Booked ahead"),
    ("gen.months36", "Generated pages", "3–6 months"),
    ("gen.flexible", "Generated pages", "Flexible"),
    ("gen.always", "Generated pages", "Always"),
    ("gen.goodToKnow", "Generated pages", "Good to know"),
    ("gen.goodCopy", "Generated pages", "The practical notes — timing, packing, the hours in between — travel with your itinerary, not a guidebook."),
    ("gen.guides", "Generated pages", "Guides"),
    ("gen.resident", "Generated pages", "Resident"),
    ("gen.routing", "Generated pages", "Routing"),
    ("gen.custom", "Generated pages", "Custom"),
    ("gen.pace", "Generated pages", "Pace"),
    ("gen.yours", "Generated pages", "Yours"),
    ("enq.journeysEnd", "Enquiry form", "Journey's end"),
    ("enq.headline1", "Enquiry form", "Get in touch,"),
    ("enq.headline2", "Enquiry form", "we'll draw the route"),
    ("enq.boardingPass", "Enquiry form", "Boarding Pass"),
    ("enq.from", "Enquiry form", "From"),
    ("enq.here", "Enquiry form", "HERE"),
    ("enq.to", "Enquiry form", "To"),
    ("enq.anywhere", "Enquiry form", "ANYWHERE"),
    ("enq.passengerName", "Enquiry form", "Passenger name"),
    ("enq.yourName", "Enquiry form", "YOUR NAME"),
    ("enq.contactEmail", "Enquiry form", "Contact email"),
    ("enq.cabinWhereTo", "Enquiry form", "Cabin — where to"),
    ("enq.flight", "Enquiry form", "Flight"),
    ("enq.gate", "Enquiry form", "Gate"),
    ("enq.open", "Enquiry form", "OPEN"),
    ("enq.seat", "Enquiry form", "Seat"),
    ("enq.class", "Enquiry form", "Class"),
    ("enq.toBeArranged", "Enquiry form", "To be arranged"),
    ("enq.confirm", "Enquiry form", "Confirm enquiry"),
    ("enq.sent", "Enquiry form", "Sent — we'll be in touch within 24 hours"),
    ("enq.cities", "Enquiry form", "London · Cape Town · Kyoto"),
]
for key, where, en in UI:
    rows.append(("ui", key, f"UI · {where}", en, "{...} placeholders stay as-is"))

# ---- prefill Arabic from the live translation docs ----
# The workbook doubles as the review sheet: whatever Arabic is currently on
# the site appears in the yellow column, so the copywriter approves or
# corrects in place and the same file re-imports in one shot.
AR = {}
for tdoc in q('*[_type=="translation" && lang=="ar"]{source, strings[]{path, value}}'):
    for s in tdoc.get("strings") or []:
        AR[(tdoc["source"], s["path"])] = s["value"]

# ------------------------------------------------------------------ xlsx
wb = Workbook()
ARIAL = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="16243C")
AR_FILL = PatternFill("solid", fgColor="FFF6D9")
BOLD_W = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BASE = Font(name=ARIAL, size=10)
NOTE = Font(name=ARIAL, size=9, color="777777")

# --- How to use ---
ws = wb.active
ws.title = "How to use"
guide = [
    ("The Retreat Collection — Arabic translation sheet", True),
    ("", False),
    ("1.  Work ONLY in the yellow 'Arabic' column on the 'Strings' sheet.", False),
    ("2.  Leave a cell empty to keep the English (e.g. brand names like 'Kulm Hotel St. Moritz').", False),
    ("3.  Never edit Ref, Field or English — they are how the upload finds its way back.", False),
    ("4.  Text in curly braces like {country} or {n} is filled by the website — keep the braces,", False),
    ("     place them where they belong in the Arabic sentence.", False),
    ("5.  '·' separators and '→' arrows can stay; numbers/coordinates stay Latin.", False),
    ("", False),
    ("Example:", True),
    ("     English:  The stays in {country}", False),
    ("     Arabic:   الإقامات في {country}", False),
    ("", False),
    ("When done, send the file back — it uploads in one shot, nothing else needed.", False),
]
for i, (text, bold) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(name=ARIAL, size=12 if i == 1 else 10, bold=bold)
ws.column_dimensions["A"].width = 110

# --- Strings ---
ws = wb.create_sheet("Strings")
headers = ["Ref", "Where it appears", "Field", "English", "Arabic", "Notes"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = BOLD_W
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"
widths = [34, 34, 30, 70, 70, 32]
for col, wdt in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = wdt

wrap = Alignment(wrap_text=True, vertical="top")
wrap_rtl = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)
for r, (doc_id, path, where, en, note) in enumerate(rows, start=2):
    ws.cell(row=r, column=1, value=f"{doc_id}|{path}").font = NOTE
    ws.cell(row=r, column=2, value=where).font = BASE
    ws.cell(row=r, column=3, value=path).font = NOTE
    c = ws.cell(row=r, column=4, value=en)
    c.font = BASE
    c.alignment = wrap
    a = ws.cell(row=r, column=5, value=AR.get((doc_id, path)))
    a.fill = AR_FILL
    a.font = Font(name=ARIAL, size=11)
    a.alignment = wrap_rtl
    ws.cell(row=r, column=6, value=note).font = NOTE

ws.auto_filter.ref = f"A1:F{len(rows)+1}"
ws.sheet_view.rightToLeft = False

out = "import/i18n/trc-arabic-translation.xlsx"
import os
os.makedirs("import/i18n", exist_ok=True)
wb.save(out)
print(f"{len(rows)} strings -> {out}")
